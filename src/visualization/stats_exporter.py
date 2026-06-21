import datetime
from pathlib import Path

import numpy as np
import pandas as pd


class ResearchExporter:
    """Multi-sheet Excel workbook for research-grade export."""

    def __init__(self, csv_rows, pred_len, analysis_results, config,
                 duration_s, video_path, track_summary=None):
        self.csv_rows = csv_rows
        self.pred_len = pred_len
        self.analysis = analysis_results
        self.config = config
        self.duration_s = duration_s
        self.video_path = str(video_path)
        self.track_summary = track_summary or []

    def save(self, path) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return False

        wb = Workbook()
        wb.remove(wb.active)

        hdr_fill = PatternFill(start_color='2D5986', end_color='2D5986', fill_type='solid')
        hdr_font = Font(color='FFFFFF', bold=True)
        alt_fill = PatternFill(start_color='EBF0F7', end_color='EBF0F7', fill_type='solid')
        yel_fill = PatternFill(start_color='FFE066', end_color='FFE066', fill_type='solid')

        def _write_df(ws, df, highlight_col_name=None, style_rows=True):
            ws.append(list(df.columns))
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal='center')
            hl_idx = None
            if highlight_col_name and highlight_col_name in df.columns:
                hl_idx = list(df.columns).index(highlight_col_name) + 1
            for i, row in enumerate(df.itertuples(index=False), start=2):
                ws.append(list(row))
                if style_rows:
                    if i % 2 == 0:
                        for cell in ws[i]:
                            cell.fill = alt_fill
                    if hl_idx:
                        ws.cell(row=i, column=hl_idx).fill = yel_fill
            if style_rows:
                for col in ws.columns:
                    w = max(len(str(cell.value or '')) for cell in col) + 2
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(w, 35)
            else:
                for idx in range(1, len(df.columns) + 1):
                    ws.column_dimensions[get_column_letter(idx)].width = 14
            ws.freeze_panes = 'A2'

        # Sheet 1: Track Summary
        if self.track_summary:
            ws = wb.create_sheet('Track Summary')
            _write_df(ws, pd.DataFrame(self.track_summary))

        # Sheet 2: Raw Detections
        if self.csv_rows:
            ws = wb.create_sheet('Raw Detections')
            _write_df(ws, pd.DataFrame(self.csv_rows), style_rows=False)

        # Sheet 2: Traffic Summary
        counting = self.analysis.get('counting_lines', [])
        if counting:
            ws = wb.create_sheet('Traffic Summary')
            rows = []
            for line in counting:
                for cls, cnt in line.get('counts', {}).items():
                    fr = line.get('flow_rates_veh_hr', {}).get(cls, 0)
                    hw = line.get('headway', {}).get(cls, {})
                    rows.append({
                        'Line': line['label'], 'Class': cls, 'Count': cnt,
                        'Flow (veh/hr)': fr,
                        'Mean Headway (s)': hw.get('mean_s', ''),
                        'Std Headway (s)': hw.get('std_s', ''),
                        'Min Headway (s)': hw.get('min_s', ''),
                        'Max Headway (s)': hw.get('max_s', ''),
                    })
            if rows:
                _write_df(ws, pd.DataFrame(rows))

        # Sheet 3: Speed Statistics (P85 highlighted — standard traffic engineering metric)
        speed = self.analysis.get('speed', {})
        if speed and speed.get('per_class'):
            ws = wb.create_sheet('Speed Statistics')
            rows = [{'Class': cls, **stats}
                    for cls, stats in speed['per_class'].items()]
            _write_df(ws, pd.DataFrame(rows), highlight_col_name='p85')

        # Sheet 4: OD Matrix
        od = self.analysis.get('od_matrix', {})
        if od and 'matrix_df' in od:
            ws = wb.create_sheet('OD Matrix')
            df = od['matrix_df'].reset_index()
            df.rename(columns={'index': 'Origin \\ Destination'}, inplace=True)
            _write_df(ws, df)

        # Sheet 5: Zone Analysis
        zone_res = self.analysis.get('zone_analysis', {})
        if zone_res and zone_res.get('zone_summaries'):
            ws = wb.create_sheet('Zone Analysis')
            rows = []
            for zs in zone_res['zone_summaries']:
                util = zs.get('utilization') or {}
                rows.append({
                    'Zone': zs['zone'],
                    'Entry Count': zs['entry_count'],
                    'Mean Dwell (s)': zs['mean_dwell_s'],
                    'Max Dwell (s)': zs['max_dwell_s'],
                    'Mean Density (p/m²)': util.get('mean_density_per_sqm', ''),
                    'Peak Density (p/m²)': util.get('peak_density_per_sqm', ''),
                    'Zone Area (m²)': util.get('zone_area_sqm', ''),
                })
            _write_df(ws, pd.DataFrame(rows))

        # Sheet 6: Time Series (zone occupancy, 1-second resolution)
        if zone_res and zone_res.get('occupancy_timeseries'):
            ws = wb.create_sheet('Time Series')
            ts = zone_res['occupancy_timeseries']
            if ts:
                max_len = max(len(v['count']) for v in ts.values())
                data = {'time_s': list(range(max_len))}
                for zname, occ in ts.items():
                    c = occ['count']
                    data[zname] = c + [0] * (max_len - len(c))
                _write_df(ws, pd.DataFrame(data))

        # Sheet 7: Congestion Events
        cong = self.analysis.get('congestion', {})
        if cong and cong.get('events'):
            ws = wb.create_sheet('Congestion Events')
            _write_df(ws, pd.DataFrame(cong['events']))

        # Sheet 8: Metadata (reproducibility)
        ws = wb.create_sheet('Metadata')
        meta = [
            ('Video', self.video_path),
            ('Duration (s)', round(self.duration_s, 1)),
            ('Processed At', datetime.datetime.now().isoformat()),
            ('YOLO Model', self.config.get('yolo_model', '')),
            ('Confidence Threshold', self.config.get('conf_thresh', '')),
            ('Sequence Length (frames)', self.config.get('seq_len', '')),
            ('Prediction Length (frames)', self.config.get('pred_len', '')),
            ('LSTM Mode', self.config.get('lstm_mode', '')),
        ]
        for row in meta:
            ws.append(row)
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 55

        wb.save(str(path))
        return True


class PlotlyExporter:
    """Interactive HTML charts and optional SVG export for publication."""

    def __init__(self, analysis_results, duration_s):
        self.analysis = analysis_results
        self.duration_s = duration_s

    def _speed_box(self):
        try:
            import plotly.graph_objects as go
        except ImportError:
            return None
        speed = self.analysis.get('speed') or {}
        track_speeds = speed.get('track_speeds', {})
        track_cls = speed.get('track_cls', {})
        if not track_speeds:
            return None
        # One sample per vehicle, not per frame — pooling raw per-frame samples
        # overweights idling/decelerating tracks and turns normal cruising
        # speeds into a wall of IQR "outlier" points. The per-vehicle sample
        # itself is each vehicle's 85th-percentile speed rather than its mean:
        # the mean is dragged down by queuing/deceleration frames and ends up
        # describing "how slow this vehicle was while stuck," not how fast it
        # travels — using the high end of its own samples (its characteristic
        # running speed) is the standard traffic-engineering convention for a
        # representative per-vehicle speed.
        by_class = {}
        for tid, speeds in track_speeds.items():
            if not speeds:
                continue
            cls = track_cls.get(tid, 'unknown')
            by_class.setdefault(cls, []).append(float(np.percentile(speeds, 85)))
        fig = go.Figure()
        for cls, ys in by_class.items():
            if len(ys) < 2:
                continue
            fig.add_trace(go.Box(name=cls, y=ys, boxmean=True, marker_color='royalblue'))
        fig.update_layout(title='Speed Distribution by Class (km/h)',
                          yaxis_title='85th-Percentile Speed per Vehicle (km/h)', template='plotly_white')
        return fig

    def _flow_timeseries(self):
        try:
            import plotly.graph_objects as go
        except ImportError:
            return None
        counting = self.analysis.get('counting_lines', [])
        if not counting:
            return None
        fig = go.Figure()
        for line in counting:
            crossings = line.get('crossings', [])
            if not crossings:
                continue
            times = [c['time_s'] for c in crossings]
            bin_s = 30
            max_t = max(times, default=0)
            n_bins = max(1, int(max_t / bin_s) + 1)
            counts = [0] * n_bins
            for t in times:
                idx = min(int(t / bin_s), n_bins - 1)
                counts[idx] += 1
            xs = [i * bin_s + bin_s / 2 for i in range(n_bins)]
            fig.add_trace(go.Scatter(x=xs, y=counts, mode='lines+markers',
                                     name=line['label']))
        fig.update_layout(title='Traffic Flow Time Series (30 s bins)',
                          xaxis_title='Time (s)',
                          yaxis_title='Crossings per 30 s',
                          template='plotly_white')
        return fig

    def _od_sankey(self):
        try:
            import plotly.graph_objects as go
        except ImportError:
            return None
        od = self.analysis.get('od_matrix', {})
        if not od or not od.get('raw'):
            return None
        zone_names = od.get('zone_names', [])
        zone_idx = {z: i for i, z in enumerate(zone_names)}
        sources, targets, values = [], [], []
        for (ori, dst), cnt in od['raw'].items():
            if ori in zone_idx and dst in zone_idx and cnt > 0:
                sources.append(zone_idx[ori])
                targets.append(zone_idx[dst])
                values.append(cnt)
        if not sources:
            return None
        fig = go.Figure(go.Sankey(
            node={'label': zone_names, 'pad': 15, 'thickness': 20},
            link={'source': sources, 'target': targets, 'value': values},
        ))
        fig.update_layout(title='Origin-Destination Flow', template='plotly_white')
        return fig

    def _zone_heatmap(self):
        try:
            import plotly.graph_objects as go
        except ImportError:
            return None
        zone_res = self.analysis.get('zone_analysis', {})
        ts = (zone_res or {}).get('occupancy_timeseries', {})
        if not ts:
            return None
        zone_names = list(ts.keys())
        max_len = max(len(v['count']) for v in ts.values())
        z_data = []
        for zn in zone_names:
            c = ts[zn]['count']
            z_data.append(c + [0] * (max_len - len(c)))
        fig = go.Figure(go.Heatmap(
            z=z_data, x=list(range(max_len)), y=zone_names,
            colorscale='Viridis', colorbar_title='Count',
        ))
        fig.update_layout(title='Zone Occupancy Over Time',
                          xaxis_title='Time (s)', yaxis_title='Zone',
                          template='plotly_white')
        return fig

    def save_html(self, path) -> bool:
        try:
            import plotly.io as pio
        except ImportError:
            return False
        named_figs = [
            ('Speed Distribution', self._speed_box()),
            ('Traffic Flow', self._flow_timeseries()),
            ('OD Flow (Sankey)', self._od_sankey()),
            ('Zone Occupancy', self._zone_heatmap()),
        ]
        named_figs = [(t, f) for t, f in named_figs if f is not None]
        if not named_figs:
            return False
        parts = ['<html><head><meta charset="utf-8">'
                 '<title>Research Analysis Charts</title></head><body>',
                 '<h1 style="font-family:sans-serif;color:#2D5986">'
                 'Research Analysis Charts</h1>']
        for title, fig in named_figs:
            parts.append(f'<h2 style="font-family:sans-serif">{title}</h2>')
            parts.append(pio.to_html(fig, full_html=False, include_plotlyjs='cdn'))
        parts.append('</body></html>')
        with open(str(path), 'w', encoding='utf-8') as f:
            f.write('\n'.join(parts))
        return True

    def save_images(self, dir_path, fmt='svg') -> bool:
        try:
            import plotly.io as pio
            import kaleido  # noqa: F401
        except ImportError:
            return False
        dir_path = Path(dir_path)
        dir_path.mkdir(parents=True, exist_ok=True)
        named_figs = {
            'speed': self._speed_box(),
            'flow': self._flow_timeseries(),
            'od': self._od_sankey(),
            'zone_occupancy': self._zone_heatmap(),
        }
        for name, fig in named_figs.items():
            if fig is None:
                continue
            try:
                pio.write_image(fig, str(dir_path / f'{name}.{fmt}'), format=fmt)
            except Exception:
                pass
        return True
